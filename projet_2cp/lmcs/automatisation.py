import requests
import openpyxl
from bs4 import BeautifulSoup
from datetime import datetime,timedelta
import time

def authors_from_excel_to_list(excel_authors_list_file):
    authors = []
    wb = openpyxl.load_workbook(excel_authors_list_file)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        nom = row[1].lower()
        nom = nom.capitalize()
        prenom = row[2]
        full_name = f"{prenom.replace(' ', '_').replace('-', '_')}_{nom.replace(' ', '_').replace('-', '_')}"
        authors.append(full_name)
    return authors

def citations_from_doi (doi):
    url = f"https://opencitations.net/index/api/v1/citation-count/{doi}"
    max_retries = 3
    retries = 0
    while retries < max_retries:
        response = requests.get(url)
        if response.status_code == 200:
            data = response.json()
            nbr_citations = int(data[0]['count'])  if data else 0
            return nbr_citations
        elif response.status_code == 429:
            print("Rate limite")
            time.sleep(100)
            retries += 1
        else:
            print("Error:", response.status_code)
            return 0
    print("Max essaies")
    return 0

def titles_from_excel_to_list(excel_publications_file):
    titles = []
    wb = openpyxl.load_workbook(excel_publications_file)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):  # Adjusted max_col to 2
        title = row[1]  # Accessing the second column (index 1)
        titles.append(title)
    return titles

def mise_a_jour_publication_to_excel( author_name, titles, excel_publications_output_file):
    publication_data = query_dblp_publication(author_name)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Chercheur","Titre Publication","Rang Chercheur", "Conf/Journal", "Année", "Type publication","lien", "Volume", "Version", "Pages", "Nombre de Citations"])
    
    for publication in publication_data:
        title = publication.get('info', {}).get('title', '')
        authors_info = publication.get('info', {}).get('authors', {}).get('author', [])
        range = None
        if isinstance(authors_info, dict):
                authors = authors_info.get('text', '')
                range=1
        else:
                authors = ', '.join([author.get('text', '') for author in authors_info])
                for i, author in enumerate(authors_info, start=1):
                    if author.get('text', '').lower() == author_name.lower().replace("_", " "):
                        range = i
                        break
        venue = publication.get('info', {}).get('venue', '')
        year = publication.get('info', {}).get('year', '')
        type_ = publication.get('info', {}).get('type', '')
        url = publication.get('info', {}).get('url', '')
        volume = publication.get('info', {}).get('volume', '')
        number = publication.get('info', {}).get('number', '')
        pages = publication.get('info', {}).get('pages', '')
        doi = publication.get('info', {}).get('doi', '')
        if citations_from_doi(doi):
            citations = citations_from_doi(doi)
        else:
            citations = 0
        if title not in titles:
            ws.append([author_name,str(title), range, str(venue), year, str(type_), url, volume, number, pages,citations])

    wb.save(excel_publications_output_file)

def mise_a_jour_publications_from_excel_to_excel(excel_authors_list_file,excel_publications_file, excel_publications_output_file):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Chercheur","Titre Publication","Rang Chercheur", "Conf/Journal", "Année", "Type publication","lien", "Volume", "Version", "Pages", "Nombre de Citations"])
    authors = authors_from_excel_to_list(excel_authors_list_file)
    titles = titles_from_excel_to_list(excel_publications_file)
    for author in authors:
        mise_a_jour_publication_to_excel(author,titles,excel_publications_output_file)
        author_wb = openpyxl.load_workbook(excel_publications_output_file)
        author_ws = author_wb.active
        for row in author_ws.iter_rows(min_row=2, values_only=True):
            if row[0] not in titles:
                ws.append(row)
    wb.save(excel_publications_output_file)

def mise_a_jour_previsionnee_from_excel_to_excel(excel_authors_list_file, excel_publications_file, excel_publications_output_file,excel_veunes_output_file,excel_authors_output_file, time_interval):
    if time_interval == '6 mois':
        interval_seconds = 6 * 30 * 24 * 60 * 60
        # interval_seconds = 30
    elif time_interval == '1 année':
        interval_seconds = 365 * 24 * 60 * 60
    elif time_interval == '18 mois':
        interval_seconds = 18 * 30 * 24 * 60 * 60
    else:
        raise ValueError("Interval invalid")

    while True:
        start_time = datetime.now()
        mise_a_jour_publications_from_excel_to_excel(excel_authors_list_file, excel_publications_file, excel_publications_output_file)
        print("Publications mises a jour")
        venues_from_excel_to_excel(excel_publications_output_file,excel_veunes_output_file)
        print("Confs_Journals mises a jour")
        init_authors_from_excel_to_excel(excel_publications_output_file,excel_authors_output_file,excel_publications_file)
        print("Profil Chercheurs mises a jour")
        end_time = start_time + timedelta(seconds=interval_seconds)
        remaining_time = (end_time - datetime.now()).total_seconds()

        # Affichage de temps restant
        while remaining_time > 0:
            print(f"Temps restant: {remaining_time:.0f} seconds")
            time.sleep(3)
            remaining_time = (end_time - datetime.now()).total_seconds()

        print("Mise a jour en cours ...")
        time.sleep(-remaining_time)


def venues_to_list(input_excel_file):
    venues = []
    wb = openpyxl.load_workbook(input_excel_file)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        venue_name = row[3]
        if venue_name not in venues:
            venues.append(venue_name)
    return venues

def query_dblp_venue(venue_name):
    url = f'https://dblp.org/search/venue/api?q={venue_name}&format=json'
    max_retries = 3
    retries = 0
    while retries < max_retries:
        response = requests.get(url)
        if response.status_code == 200:
            data = response.json()
            return data.get('result', {}).get('hits', {}).get('hit', [])
        elif response.status_code == 429:
            print("Rate limite")
            time.sleep(100)
            retries += 1
        else:
            print("Error:", response.status_code)
            return []
    print("Max essaies")
    return []

def venue_to_excel(venue_name, excel_output_file):
    venue_info = query_dblp_venue(venue_name)
    if venue_info:
        info = venue_info[0].get('info', {})
        acronym = info.get('acronym', '')
        venue_type = info.get('type', '')
        url = info.get('url', '')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Conf/Journal", "Acronyme", "Type", "Lien"])
        ws.append([venue_name, acronym, venue_type, url])
        wb.save(excel_output_file)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Conf/Journal", "Acronyme", "Type", "Lien"])
        ws.append([venue_name, venue_name, '', ''])
        wb.save(excel_output_file)
        print(f"Venue '{venue_name}' not found on DBLP.")

def venues_from_excel_to_excel(input_excel_file, output_excel_file):
    venues = venues_to_list(input_excel_file)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Conf/Journal", "Acronyme", "Type", "Lien"])

    for venue_name in venues:
        venue_to_excel(venue_name, output_excel_file)
        venue_wb = openpyxl.load_workbook(output_excel_file)
        venue_ws = venue_wb.active
        for row in venue_ws.iter_rows(min_row=2, values_only=True):
            ws.append(row)

    wb.save(output_excel_file)


def query_dblp_author(author_name):
    url = f'https://dblp.org/search/author/api?q={author_name}&format=json'
    max_retries = 3
    retries = 0
    while retries < max_retries:
        response = requests.get(url)
        if response.status_code == 200:
            data = response.json()
            return data.get('result', {}).get('hits', {}).get('hit', [])
        elif response.status_code == 429:
            print("Rate limite")
            time.sleep(100)
            retries += 1
        else:
            print("Error:", response.status_code)
            return []
    print("Max essaies")
    return []

def orcid(content):
    soup = BeautifulSoup(content, 'html.parser')
    orcid_element = soup.find('li', class_='orcid drop-down')
    if orcid_element:
        orcid_link = orcid_element.find('a', {'href': lambda x: x and 'orcid.org' in x})
        if orcid_link:
            orcid_url = orcid_link['href']
            orcid_id = orcid_url.split('/')[-1]

            return orcid_id
    else:
        return None
    
def lien_google_scholar(content):
    soup = BeautifulSoup(content, 'html.parser')
    search_element = soup.find('li', class_='search drop-down')
    if search_element:
        google_scholar_link = search_element.find('a', href=lambda x: x and 'scholar.google.com' in x)
        if google_scholar_link:
            return google_scholar_link['href']
    return None
    

def query_dblp_publication(author_name):
    url = f'https://dblp.org/search/publ/api?q=author%3A{author_name}%3A&h=1000&format=json'
    max_retries = 3
    retries = 0
    while retries < max_retries:
        response = requests.get(url)
        if response.status_code == 200:
            data = response.json()
            return data.get('result', {}).get('hits', {}).get('hit', [])
        elif response.status_code == 429:
            print("Rate limite")
            time.sleep(100)
            retries += 1
        else:
            print("Error:", response.status_code)
            return []
    print("Max essaies")
    return []

def init_publication_to_excel(author_name, excel_publications_file):
    publication_data = query_dblp_publication(author_name)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Chercheur","Titre Publication","Rang Chercheur", "Conf/Journal", "Année", "Type publication","lien", "Volume", "Version", "Pages", "Nombre de Citations"])
    
    for publication in publication_data:
        title = publication.get('info', {}).get('title', '')
        authors_info = publication.get('info', {}).get('authors', {}).get('author', [])
        range = None
        if isinstance(authors_info, dict):
                authors = authors_info.get('text', '')
                range=1
        else:
                authors = ', '.join([author.get('text', '') for author in authors_info])
                for i, author in enumerate(authors_info, start=1):
                    if author.get('text', '').lower() == author_name.lower().replace("_", " "):
                        range = i
                        break
        venue = publication.get('info', {}).get('venue', '')
        year = publication.get('info', {}).get('year', '')
        type_ = publication.get('info', {}).get('type', '')
        url = publication.get('info', {}).get('url', '')
        volume = publication.get('info', {}).get('volume', '')
        number = publication.get('info', {}).get('number', '')
        pages = publication.get('info', {}).get('pages', '')
        doi = publication.get('info', {}).get('doi', '')
        if citations_from_doi(doi):
            citations = citations_from_doi(doi)
        else:
            citations = 0
        ws.append([author_name,str(title),  range , str(venue), year, str(type_), url, volume, number, pages,citations])

    wb.save(excel_publications_file)

def get_author_content(author_url):
    max_retries = 3
    retries = 0
    while retries < max_retries:
        response = requests.get(author_url)
        if response.status_code == 200:
            return response.content
        elif response.status_code == 429:
            print("Rate limite")
            time.sleep(100)
            retries += 1
        else:
            print("Error:", response.status_code)
            return None
    print("Max essaies")
    return None

def init_author_to_excel(author_name,publications_excel_file, excel_output_file):
    author_info = query_dblp_author(author_name)
    wb_author = openpyxl.Workbook()
    ws_author = wb_author.active
    ws_author.append(["ORCID", "Lien DBLP","Lien Google Scholar","Lien Research Gate","h-index"])

    if author_info:
        info = author_info[0].get('info', {})
        author_url = info.get('url', '')
        content = get_author_content(author_url)
        author_orcid = orcid(content)
        author_google_scholar = lien_google_scholar(content)
        hindex = hindex_to_excel(publications_excel_file,author_name)
        print(hindex)
        author_researchgate = f'https://www.researchgate.net/search/researcher?q={author_name}'
        ws_author.append([author_orcid, author_url,author_google_scholar,author_researchgate,hindex])
    else:
        ws_author.append(['', '', '','',0])
        print(f"Author '{author_name}' not found on DBLP.")
    wb_author.save(excel_output_file)

def hindex_to_excel(publications_excel_file, author_name):
    citations = get_author_citations(publications_excel_file, author_name)
    print(citations)
    if citations:
        h_index = calculate_h_index(citations)
        if h_index is not None:
            return h_index

def get_author_citations(publications_excel_file, author_name):
    citations = []
    wb = openpyxl.load_workbook(publications_excel_file)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if author_name in row[0]:
            citations.append(row[10])
    return citations

def calculate_h_index(citations):
    citations = sorted(citations, reverse=True)
    for index, value in enumerate(citations):
        if index + 1 > value:
            return index
    return len(citations)


def init_publications_from_excel_to_excel(excel_authors_list_file, excel_output_file):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Clé Chercheur","Titre Publication","Rang Chercheur", "Conf/Journal", "Année", "Type publication","Lien", "Volume", "Version", "Pages", "Nombre de Citations"])
    authors = authors_from_excel_to_list(excel_authors_list_file)
    print(authors)
    for author in authors:
        init_publication_to_excel(author, excel_output_file)
        author_wb = openpyxl.load_workbook(excel_output_file)
        author_ws = author_wb.active

        for row in author_ws.iter_rows(min_row=2, values_only=True):
            ws.append(row)
    wb.save(excel_output_file)

def init_authors_from_excel_to_excel(input_excel_file, output_excel_file,publications_excel_file):
    wb_input = openpyxl.load_workbook(input_excel_file)
    ws_input = wb_input.active

    wb_output = openpyxl.Workbook()
    ws_output = wb_output.active
    headers = [cell.value for cell in ws_input[1]] + ["ORCID", "Lien DBLP", "Lien Google Scholar", "Lien Research Gate", "h-index"]
    ws_output.append(headers)
    for row in ws_input.iter_rows(min_row=2, values_only=True):
        author_info = dict(zip(headers[:len(row)], row))
        nom = row[1].lower()
        nom = nom.capitalize()
        prenom = row[2]
        author_name = f"{prenom.replace(' ', '_').replace('-', '_')}_{nom.replace(' ', '_').replace('-', '_')}"
        init_author_to_excel(author_name, publications_excel_file, output_excel_file)
        author_wb = openpyxl.load_workbook(output_excel_file)
        author_ws = author_wb.active
        if author_ws.max_row > 1:
            author_row_values = [cell.value for cell in author_ws[2]]
            ws_output.append(tuple(row) + tuple(author_row_values))
        else:
            ws_output.append(row + [''] * 5)
    wb_output.save(output_excel_file)

def projets_to_excel(output_excel_file):
    url=f'https://lmcs.esi.dz/recherche/projets/projets-nationaux/projets-prfu-en-cours/'
    max_retries = 3
    retries = 0
    while retries < max_retries:
        response = requests.get(url)
        if response.status_code == 200:
            soup = BeautifulSoup(response.content, 'html.parser')
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["title", "code", "responsible", "members"])

            for project_info in soup.find_all('p'):
                project_data = {}

                title_tag = project_info.find('strong', string='Projet')
                if title_tag:
                    project_data['title'] = title_tag.next_sibling.strip()
                    
                    code_tag = project_info.find('strong', string='Code')
                    if code_tag:
                        project_data['code'] = code_tag.next_sibling.strip()
                    
                    responsible_tag = project_info.find('strong', string='Responsable')
                    if responsible_tag:
                        project_data['responsible'] = responsible_tag.next_sibling.strip()

                    members_tag = project_info.find_next_sibling('table')
                    if members_tag:
                        members = members_tag.find_all('td', width="247")
                        project_data['members'] = [member.text.strip() for member in members]

                    ws.append([project_data.get('title', ''),
                            project_data.get('code', ''),
                            project_data.get('responsible', ''),
                            ','.join(project_data.get('members', []))])
                
            wb.save(output_excel_file)
        elif response.status_code == 429:
            print("Rate limite")
            time.sleep(100)
            retries += 1
        else:
            print("Error:", response.status_code)
            return None
    print("Max essaies")
    return None

# init_publications_from_excel_to_excel("authors_list.xlsx", "init_publications.xlsx")
# mise_a_jour_publications_from_excel_to_excel("authors_list.xlsx","init_publications.xlsx","mise_a_jour_publications.xlsx")
# mise_a_jour_previsionnee_from_excel_to_excel("authors_list.xlsx","init_publications.xlsx","mise_a_jour_publications.xlsx","6 mois")
# venues_from_excel_to_excel("init_publications.xlsx","venus.xlsx")
# init_authors_from_excel_to_excel("authors_list.xlsx", "authors.xlsx","init_publications.xlsx")
# projets_to_excel("projets.xlsx")
# init_author_to_excel("Fahima Nader","init_publications.xlsx", "authors.xlsx")